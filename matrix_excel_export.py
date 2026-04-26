"""
Dán TSV bảng quyết định (matrix 4+ cột mã) → .xlsx định dạng dự án:
Tahoma 8, nền xanh đậm cột mục + hàng mã, merge A/B/C theo cấu trúc, dropdown O ở vùng tick.

Ô TSV bắt đầu bằng = được ghi đúng công thức Excel (ví dụ =COUNTIF(E10:HG10,"P"), =SUM(N8,-A8,-C8)).
Công thức hàng 6 (Passed/Failed/…) có thể ghi đè bằng biến môi trường: MATRIX_F6_PASSED, MATRIX_F6_FAILED,
MATRIX_F6_UNTESTED, MATRIX_F6_N, MATRIX_F6_A, MATRIX_F6_B, MATRIX_F6_TOTAL.
"""
from __future__ import annotations

import io
import os
import re
from typing import List, Optional, Tuple

from excel_testplan_to_table import parse_tsv

# Nền xanh (dark blue 26 / #002060) — aRGB
FILL_DARK = "002060"
FONT_MAIN = "Tahoma"
FC_WHITE = "FFFFFFFF"
FC_BLACK = "FF000000"

_CASE_HEADER = re.compile(r"^(?:AUTH|ADM|STU|TCH)-\d{3}\s*$", re.IGNORECASE)
_MAIN_A = re.compile(r"^(?:Condition|Confirm|Result)$", re.IGNORECASE)
_EXACT_B = re.compile(
    r"^(?:Precondition|Return|Exception|Log message)$",
    re.IGNORECASE,
)

_DATE_PAT = re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\b")

# Dòng cấp 1: metadata + 2 dòng tổng hợp công thức + 1 dòng trống; ma trận bắt đầu từ 1+REPORT_HEADER_ROWS
REPORT_HEADER_ROWS = 7
# Có thể ghi đè bằng biến môi trường: MATRIX_DEFAULT_FUNCTION_CODE, etc.
_DEFAULT_FUNCTION_CODE = os.environ.get("MATRIX_DEFAULT_FUNCTION_CODE", "UTC001")
_DEFAULT_FUNCTION_NAME = os.environ.get("MATRIX_DEFAULT_FUNCTION_NAME", "Login")
_DEFAULT_PERSON = os.environ.get("MATRIX_DEFAULT_PERSON", "Phạm Đăng Khôi")
_DEFAULT_LINES_OF_CODE = os.environ.get("MATRIX_DEFAULT_LINES_OF_CODE", "160")


def _is_excel_formula_str(s: str) -> bool:
    t = (s or "").strip()
    return bool(t) and t.startswith("=")


def _set_openpyxl_cell_value(cell, v: str | int | float | None) -> None:
    """Gán giá trị ô: chuỗi bắt đầu = là công thức Excel (đúng dạng người dùng nhập)."""
    from openpyxl.cell.cell import MergedCell

    if isinstance(cell, MergedCell):
        return
    if v is None:
        cell.value = None
        return
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        cell.value = v
        return
    s = str(v)
    st = s.strip()
    if not st:
        cell.value = None
        return
    if st.startswith("="):
        cell.value = st
    else:
        cell.value = s


def _row6_formula_env_overrides() -> dict[int, str | None]:
    """Công thức tùy chọn hàng 6 (Passed / Failed / …) — cùng cú pháp Excel, ví dụ =COUNTIF(E43:HG43,'P')."""
    def g(k: str) -> str | None:
        t = (os.environ.get(k) or "").strip()
        return t if t else None

    return {
        1: g("MATRIX_F6_PASSED"),
        3: g("MATRIX_F6_FAILED"),
        5: g("MATRIX_F6_UNTESTED"),
        11: g("MATRIX_F6_N"),
        12: g("MATRIX_F6_A"),
        13: g("MATRIX_F6_B"),
        14: g("MATRIX_F6_TOTAL"),
    }


def _last_date_in_cell(s: str) -> str:
    """Executed Date: chỉ lấy ngày cuối (sau mũi tên hoặc bản ghi date cuối cùng)."""
    t = (s or "").replace("\n", " ").replace("\r", " ").strip()
    if not t:
        return ""
    for sep in ("→", "->", "=>"):
        if sep in t:
            t = t.split(sep)[-1].strip()
    dates = _DATE_PAT.findall(t)
    if dates:
        return dates[-1]
    # sau dấu ; lấy đoạn có date, hoặc cả đoạn cuối
    parts = re.split(r"[;]", t)
    for p in reversed(parts):
        p = p.strip()
        m = _DATE_PAT.search(p)
        if m:
            return m.group(0)
    return t


def _slash_right_label(s: str) -> str:
    """Failed/Passed -> Passed (lấy phần sau dấu / nếu đúng dạng A/B)."""
    t = (s or "").replace("\n", " ").replace("\r", " ").strip()
    if not t or "/" not in t:
        return t
    if t.startswith("http"):
        return t
    if (t.count("/") == 1) and "http" not in t:
        a, b = t.split("/", 1)
        if 0 < len(a) < 40 and 0 < len(b) < 40:
            return b.strip() or t
    return t


def _is_executed_date_row_b(b: str) -> bool:
    b = (b or "").strip()
    if not b:
        return False
    if b == "Executed Date" or b.startswith("Executed Date"):
        return True
    return b.startswith("Executed") and "Date" in b


def _is_passed_failed_row_b(b: str) -> bool:
    b = (b or "").strip().lower().replace(" ", "")
    if not b:
        return False
    return b == "passed/failed" or b.startswith("passed/failed")


def _is_type_nab_block_label(s: str) -> bool:
    t = (s or "").replace("\n", " ").replace("\r", " ").strip()
    if not t:
        return False
    if t.startswith("Type (") and "Normal" in t and "Abnormal" in t:
        return True
    return "N:" in t and "Normal" in t and "Abnormal" in t and "Boundary" in t


def _clear_bcd_type_from_result_row_if_same_line(grid: List[List[str]], r_res: int) -> None:
    """Bỏ chữ dài 'Type (N: Normal, …' ở B–D trên cùng hàng A=Result (không ngang hàng với 'Result')."""
    if r_res < 0 or r_res >= len(grid):
        return
    if _row_i_col0_strips(grid, r_res) != "Result":
        return
    row = grid[r_res]
    for ci in (1, 2, 3):
        if len(row) > ci and _is_type_nab_block_label(str(row[ci] or "")):
            row[ci] = ""


def _find_r_type_row(grid: List[List[str]], nrows: int) -> int:
    r1 = next(
        (
            i
            for i in range(nrows)
            if "Type" in _row_i_col1_strips(grid, i) and "Boundary" in (grid[i][1] or "")
        ),
        -1,
    )
    if r1 >= 0:
        return r1
    return next(
        (i for i in range(nrows) if _row_i_col1_strips(grid, i).startswith("Type (")),
        -1,
    )


def _find_r_passed_failed_row(grid: List[List[str]], nrows: int) -> int:
    for i in range(nrows):
        if _is_passed_failed_row_b(_row_i_col1_strips(grid, i)):
            return i
    return -1


def _find_nab_data_row(
    grid: List[List[str]], nrows: int, c0: int, c1: int, r_type: int, r_res: int
) -> int:
    """Dòng lưới có ít nhất một ô cột mã = N, A, hoặc B (dòng mẫu N/A/B)."""
    if c0 < 0 or c1 < c0:
        return -1

    def score_nab(ri: int) -> int:
        if ri < 0 or ri >= len(grid):
            return 0
        r = grid[ri]
        s = 0
        for ci in range(c0, c1 + 1):
            if ci < len(r) and (r[ci] or "").strip() in ("N", "A", "B"):
                s += 1
        return s

    for ri in (r_type, r_type + 1) if r_type >= 0 else ():
        if 0 <= ri < nrows and score_nab(ri) > 0:
            return ri
    start = r_res + 1 if r_res >= 0 else 0
    for ri in range(start, nrows):
        if score_nab(ri) > 0:
            return ri
    return -1


def _row_excel(ri: int, offset: int) -> int:
    """Grid 0-based row → Excel 1-based row, sau vùng header báo cáo."""
    return ri + 1 + offset


def _write_unit_test_metadata_header(
    ws,
    offset: int,
    c0: int,
    c1: int,
    h_row: int,
    nab_row: int,
    r_pf: int,
    ncols_grid: int,
) -> None:
    """
    Nhãn (A–B, E–J 1–3, tổng hợp 5) căn trái, in đậm, viết thường; hàng 5 gộp K–M → "N/A/B" (giữa).
    Giá trị C–D / N (italic) giữ căn giữa như bảng mẫu.
    """
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    last_c = 16
    _ = ncols_grid

    thin = Side(style="thin", color="000000")
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_8 = Font(name=FONT_MAIN, size=8, color=FC_BLACK, bold=False)
    font_b = Font(name=FONT_MAIN, size=8, color=FC_BLACK, bold=True)
    font_i = Font(name=FONT_MAIN, size=8, color=FC_BLACK, bold=False, italic=True)
    a_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def sc(r: int, c: int, v: str | int | None = None):
        cl = ws.cell(row=r, column=c)
        # MergedCell: chỉ dùng khi tô border / font, không gán value (read-only); v=None bỏ qua gán.
        if v is not None:
            _set_openpyxl_cell_value(cl, v)
        return cl

    def border_block(r0: int, c0: int, r1: int, c1: int) -> None:
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                sc(r, c).border = b_all

    def apply_meta_14_merges(r: int) -> None:
        """Dòng 1–4: gộp A–B, C–D, E–J; dòng 1–3 gộp K–P; dòng 4 gộp N–P (K–M tách)."""
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
        if r in (1, 2, 3):
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=16)
        else:
            ws.merge_cells(start_row=r, start_column=14, end_row=r, end_column=16)

    def apply_row56_merges(r: int) -> None:
        """Dòng 5–6: A–B, C–D, E–J, N–P; K, L, M riêng (11–13)."""
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
        ws.merge_cells(start_row=r, start_column=14, end_row=r, end_column=16)

    for r in (1, 2, 3, 4):
        apply_meta_14_merges(r)
    for r in (5, 6):
        apply_row56_merges(r)
    # Một ô tiêu đề N/A/B (gộp K–M) chỉ ở hàng 5
    ws.merge_cells(start_row=5, start_column=11, end_row=5, end_column=13)

    # Nhãn cột A–B: viết thường, bold (đúng khoảng "Lines  of code")
    _LBL_FN_CODE = "Function Code"
    _LBL_CREATED = "Created By"
    _LBL_LOC = "Lines  of code"
    _LBL_TESTREQ = "Test requirement"

    # Hàng 1: A:B nhãn; C:D UTC; E:J Function name; K:P Login
    sc(1, 1, _LBL_FN_CODE)
    sc(1, 1).font = font_b
    sc(1, 1).alignment = a_left
    sc(1, 3, _DEFAULT_FUNCTION_CODE)
    sc(1, 3).font = font_i
    sc(1, 3).alignment = a_c
    sc(1, 5, "Function Name")
    sc(1, 5).font = font_b
    sc(1, 5).alignment = a_left
    sc(1, 11, _DEFAULT_FUNCTION_NAME)
    sc(1, 11).font = font_i
    sc(1, 11).alignment = a_c

    # Hàng 2
    sc(2, 1, _LBL_CREATED)
    sc(2, 1).font = font_b
    sc(2, 1).alignment = a_left
    sc(2, 3, _DEFAULT_PERSON)
    sc(2, 3).font = font_i
    sc(2, 3).alignment = a_c
    sc(2, 5, "Executed By")
    sc(2, 5).font = font_b
    sc(2, 5).alignment = a_left
    sc(2, 11, _DEFAULT_PERSON)
    sc(2, 11).font = font_i
    sc(2, 11).alignment = a_c

    # Hàng 3: C:D 160; E:J Lack of test cases; K:P trống
    sc(3, 1, _LBL_LOC)
    sc(3, 1).font = font_b
    sc(3, 1).alignment = a_left
    sc(3, 3, _DEFAULT_LINES_OF_CODE)
    sc(3, 3).font = font_i
    sc(3, 3).alignment = a_c
    sc(3, 5, "Lack of test cases")
    sc(3, 5).font = font_b
    sc(3, 5).alignment = a_left
    sc(3, 11, "")

    # Hàng 4: A:B nhãn, còn lại rỗng
    sc(4, 1, _LBL_TESTREQ)
    sc(4, 1).font = font_b
    sc(4, 1).alignment = a_left
    for c in (3, 5, 11, 12, 13, 14):
        sc(4, c, "")

    # Hàng 5: viết thường, đậm, căn trái; K–M gộp 1 ô "N/A/B" căn giữa
    for col, t in ((1, "Passed"), (3, "Failed"), (5, "Untested")):
        sc(5, col, t)
        sc(5, col).font = font_b
        sc(5, col).alignment = a_left
    sc(5, 11, "N/A/B")
    sc(5, 11).font = font_b
    sc(5, 11).alignment = a_c
    sc(5, 14, "Total test cases")
    sc(5, 14).font = font_b
    sc(5, 14).alignment = a_left

    use_formulas = c0 >= 0 and c1 >= c0 and h_row >= 0 and nab_row >= 0 and r_pf >= 0
    f6e = _row6_formula_env_overrides()
    if use_formulas:
        cL = get_column_letter(c0 + 1)
        cR = get_column_letter(c1 + 1)
        h_e = h_row + 1 + offset
        t_e = nab_row + 1 + offset
        pf_e = r_pf + 1 + offset
        rng_t = f"${cL}${t_e}:${cR}${t_e}"
        rng_pf = f"${cL}${pf_e}:${cR}${pf_e}"
        h_hdr = f"${cL}${h_e}:${cR}${h_e}"
        sc(6, 1, f6e[1] or f"=COUNTIF({rng_pf},\"P\")")
        sc(6, 1).font = font_8
        sc(6, 1).alignment = a_c
        sc(6, 3, f6e[3] or f"=COUNTIF({rng_pf},\"F\")+COUNTIF({rng_pf},\"Failed\")")
        sc(6, 3).font = font_8
        sc(6, 3).alignment = a_c
        sc(6, 5, f6e[5] or f"=COUNTBLANK({rng_pf})")
        sc(6, 5).font = font_8
        sc(6, 5).alignment = a_c
        sc(6, 11, f6e[11] or f"=COUNTIF({rng_t},\"N\")")
        sc(6, 12, f6e[12] or f"=COUNTIF({rng_t},\"A\")")
        sc(6, 13, f6e[13] or f"=COUNTIF({rng_t},\"B\")")
        for c in (11, 12, 13):
            sc(6, c).font = font_8
            sc(6, c).alignment = a_c
        sc(6, 14, f6e[14] or f"=COLUMNS({h_hdr})")
        sc(6, 14).font = font_8
        sc(6, 14).alignment = a_c
    else:
        for col in (1, 3, 5, 11, 12, 13, 14):
            val = f6e.get(col)
            c = sc(6, col, val if val is not None else "")
            c.font = font_8
            c.alignment = a_c
    for c in range(1, last_c + 1):
        c7 = sc(7, c, None)
        c7.font = font_8
    # Chỉ kẻ A1:P6 (bảng metadata); hàng 7 trống — không kẻ.
    border_block(1, 1, 6, last_c)


def _join_b_c_d_into_b(grid: List[List[str]], nrows: int, h_row: Optional[int]) -> None:
    """Trước khi merge B:D: gom nội dung B,C,D vào ô B (tránh mất chữ khi chỉ C/D có mô tả). Bỏ qua hàng mã TCH (sẽ gộp A1:D1)."""
    for ri in range(nrows):
        if h_row is not None and ri == h_row:
            continue
        if ri >= len(grid):
            continue
        r = grid[ri]
        while len(r) < 4:
            r.append("")
        b = (r[1] or "").strip()
        c = (r[2] or "").strip()
        d = (r[3] or "").strip()
        if _is_excel_formula_str(b) or _is_excel_formula_str(c) or _is_excel_formula_str(d):
            continue
        if not b and not c and not d:
            continue
        parts = [p for p in (b, c, d) if p]
        if len(parts) == 1 and parts[0] == b and b:
            continue
        r[1] = " ".join(parts)
        r[2] = r[3] = ""


def _preprocess_data_columns(
    grid: List[List[str]],
    nrows: int,
    c0: int,
    c1: int,
    r_res: int,
) -> None:
    """Cột mã: ngày cuối, Passed/Failed lấy phần sau /, xóa 'Result' thừa ở B–D hàng bắt đầu khối Result."""
    use_cases = 0 <= c0 <= c1
    for ri in range(nrows):
        b = _row_i_col1_strips(grid, ri)
        if not b:
            continue
        row = grid[ri] if ri < len(grid) else []
        w = len(row)
        if use_cases and _is_executed_date_row_b(b):
            for ci in range(c0, c1 + 1):
                if ci < w and (row[ci] or "").strip():
                    if _is_excel_formula_str(str(row[ci] or "")):
                        continue
                    row[ci] = _last_date_in_cell(str(row[ci] or ""))
        if use_cases and _is_passed_failed_row_b(b):
            for ci in range(c0, c1 + 1):
                if ci < w and (row[ci] or "").strip():
                    if _is_excel_formula_str(str(row[ci] or "")):
                        continue
                    row[ci] = _slash_right_label(str(row[ci] or ""))
    if 0 <= r_res < nrows and r_res < len(grid):
        row = grid[r_res]
        for ci in (1, 2, 3):
            if len(row) > ci and (row[ci] or "").strip() == "Result":
                row[ci] = ""
    _clear_bcd_type_from_result_row_if_same_line(grid, r_res)


def _pad_grid(grid: List[List[str]]) -> Tuple[List[List[str]], int, int]:
    if not grid:
        return [[]], 0, 0
    w = max(len(r) for r in grid)
    out = []
    for r in grid:
        row = [r[i] if i < len(r) else "" for i in range(w)]
        out.append(row)
    return out, len(out), w


def _find_code_header_row(grid: List[List[str]]) -> Tuple[Optional[int], int, int]:
    for ri, r in enumerate(grid):
        for ci, c in enumerate(r):
            if c and _CASE_HEADER.match(c.strip()):
                c0 = ci
                c1 = ci
                for cj in range(ci + 1, len(r)):
                    if r[cj] and _CASE_HEADER.match(r[cj].strip()):
                        c1 = cj
                    else:
                        break
                return ri, c0, c1
    return None, 0, 0


def _row_i_col0_strips(grid: List[List[str]], i: int) -> str:
    r = grid[i]
    return (r[0] or "").strip() if r else ""


def _row_i_col1_strips(grid: List[List[str]], i: int) -> str:
    r = grid[i]
    if not r or len(r) < 2:
        return ""
    return (r[1] or "").strip()


def _row_all_empty(r: List[str]) -> bool:
    return not r or not any((c or "").strip() for c in r)


def _collapse_blank_line_between_result_and_type(grid: List[List[str]], r_res: int, r_type: int) -> Tuple[int, int]:
    """Bỏ hàng hoàn toàn rỗng ngay dưới hàng A=Result, trước hàng Type (tránh dòng B/C/D trống thừa). Trả (r_type mới, nrows mới) sau 1 bước."""
    nrows = len(grid)
    if r_res < 0 or r_type < 0 or r_type != r_res + 2:
        return r_type, nrows
    if r_res + 1 >= nrows:
        return r_type, nrows
    if not _row_all_empty(grid[r_res + 1]):
        return r_type, nrows
    del grid[r_res + 1]
    nrows2 = nrows - 1
    return r_type - 1, nrows2


def grid_to_workbook_bytes(grid: List[List[str]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("Cần cài: pip install openpyxl") from e

    grid, nrows, ncols = _pad_grid(grid)
    if nrows < 1 or ncols < 1:
        raise ValueError("Bảng trống.")

    h_row, c0, c1 = _find_code_header_row(grid)
    r_cond = next((i for i in range(nrows) if _row_i_col0_strips(grid, i) == "Condition"), -1)
    r_conf = next((i for i in range(nrows) if _row_i_col0_strips(grid, i) == "Confirm"), -1)
    r_res = next((i for i in range(nrows) if _row_i_col0_strips(grid, i) == "Result"), -1)
    if r_res < 0:
        r_res = next((i for i in range(nrows) if _row_i_col1_strips(grid, i) == "Result"), -1)
    r_type = _find_r_type_row(grid, nrows)
    r_type, nrows = _collapse_blank_line_between_result_and_type(grid, r_res, r_type)
    grid, nrows, ncols = _pad_grid(grid)
    if h_row is not None and 0 <= c0 <= c1:
        _preprocess_data_columns(grid, nrows, c0, c1, r_res)
    else:
        _preprocess_data_columns(grid, nrows, -1, -1, r_res)
    # Sau khi xóa "Type" trùng ở cùng hàng Result, tìm lại dòng Type cho merge B–D
    r_type = _find_r_type_row(grid, nrows)
    r_pf = _find_r_passed_failed_row(grid, nrows)
    _join_b_c_d_into_b(grid, nrows, h_row)
    nab_r = _find_nab_data_row(grid, nrows, c0, c1, r_type, r_res)

    T = REPORT_HEADER_ROWS
    h_meta = h_row if h_row is not None else -1
    c0m, c1m = (c0, c1) if h_row is not None else (-1, -1)

    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover
        raise RuntimeError("no active sheet")
    ws.title = "Matrix"

    _write_unit_test_metadata_header(ws, T, c0m, c1m, h_meta, nab_r, r_pf, ncols)

    for ri, r in enumerate(grid):
        wrow = _row_excel(ri, T)
        for ci, v in enumerate(r):
            c = ws.cell(row=wrow, column=ci + 1)
            _set_openpyxl_cell_value(c, v if v or _is_excel_formula_str(str(v or "")) else None)
            c.font = Font(name=FONT_MAIN, size=8, color=FC_BLACK, bold=False)
            c.alignment = Alignment(vertical="center", wrap_text=True)

    def merge_1a(r1: int, c1: int, r2: int, c2: int) -> None:
        if r1 > r2 or c1 > c2:
            return
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    # Cột A: Condition
    if r_cond >= 0 and r_conf > r_cond:
        merge_1a(r_cond + 1 + T, 1, r_conf + T, 1)
    # Cột A: Confirm → hết dòng trước hàng bắt đầu khối Result (A = Result)
    if r_conf >= 0 and r_res > r_conf:
        merge_1a(r_conf + 1 + T, 1, r_res + T, 1)
    # Cột A: Result (A=Result) + … (trước Type)
    if r_res >= 0:
        a_val = _row_i_col0_strips(grid, r_res) or "Result"
        ws.cell(row=r_res + 1 + T, column=1, value=a_val)
        if r_type >= 0 and r_type > r_res and r_type + 3 < nrows:
            end_1b = min(r_type + 4, nrows)
            merge_1a(r_res + 1 + T, 1, end_1b + T, 1)
        else:
            merge_1a(r_res + 1 + T, 1, min(r_res + 5, nrows) + T, 1)
    # B:D mỗi hàng ma trận (trừ hàng mã TCH: gộp A–D ở bước sau)
    if ncols >= 4:
        for ri in range(nrows):
            if h_row is not None and ri == h_row:
                continue
            merge_1a(ri + 1 + T, 2, ri + 1 + T, 4)
    # A:D hàng mã thử
    if h_row is not None and nrows > h_row and ncols >= 4:
        merge_1a(h_row + 1 + T, 1, h_row + 1 + T, 4)

    if h_row is not None and c0 <= c1 and c1 >= 0:
        c_start = c0 + 1
        c_end = c1 + 1
        if r_type > h_row + 1:
            last_0 = r_type - 1
        elif r_type >= 0:
            last_0 = max(h_row, r_type - 1)
        else:
            last_0 = nrows - 5 if nrows > h_row + 5 else nrows - 1
        if last_0 > h_row:
            a1 = f"{get_column_letter(c_start)}{h_row + 2 + T}"
            a2 = f"{get_column_letter(c_end)}{last_0 + 1 + T}"
            dv = DataValidation(
                type="list",
                formula1='"O"',
                allow_blank=True,
                showErrorMessage=True,
            )
            ws.add_data_validation(dv)
            dv.add(f"{a1}:{a2}")

    fill = PatternFill(start_color=FILL_DARK, end_color=FILL_DARK, fill_type="solid")
    font_w = Font(name=FONT_MAIN, size=8, bold=True, color=FC_WHITE)
    a_align_main = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for ri in range(nrows):
        a0 = _row_i_col0_strips(grid, ri)
        if a0 in ("Condition", "Confirm", "Result") or (a0 and _MAIN_A.match(a0)):
            cell = ws.cell(row=ri + 1 + T, column=1)
            if not cell.value and a0:
                cell.value = a0
            cell.font = font_w
            cell.fill = fill
            cell.alignment = a_align_main
    for ri in range(nrows):
        b = _row_i_col1_strips(grid, ri)
        if b in ("Precondition", "Return", "Exception", "Log message") or (b and _EXACT_B.match(b)):
            c = ws.cell(row=ri + 1 + T, column=2)
            c.font = Font(name=FONT_MAIN, size=8, bold=True, color=FC_BLACK)
    if h_row is not None and 0 <= c0 <= c1:
        corner = ws.cell(row=h_row + 1 + T, column=1)
        corner.fill = fill
        corner.font = font_w
        corner.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=h_row + 1 + T, column=c + 1)
            cell.font = Font(name=FONT_MAIN, size=8, bold=True, color=FC_WHITE)
            cell.fill = PatternFill(start_color=FILL_DARK, end_color=FILL_DARK, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if r_res >= 0:
        cr = ws.cell(row=r_res + 1 + T, column=1)
        if not (cr.value and str(cr.value).strip()):
            cr.value = "Result"
        cr.font = font_w
        cr.fill = fill
        cr.alignment = a_align_main

    # Chỉ kẻ bảng ma trận từ hàng 8 (T+1) tới hết dữ liệu; cột 1..ncols. Ô còn lại không kẻ.
    thin2 = Side(style="thin", color="000000")
    b_all2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
    r2_start = T + 1
    r2_end = nrows + T
    for r in range(r2_start, r2_end + 1):
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = b_all2

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def tsv_to_xlsx_bytes(text: str) -> bytes:
    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    if not t.strip():
        raise ValueError("Dữ liệu trống.")
    g = parse_tsv(t)
    if not g:
        raise ValueError("Không parse được dòng nào.")
    return grid_to_workbook_bytes(g)
